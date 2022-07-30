### "Pseudo codigo"
# Introducir todos los posibles valores de nxm
# Por cada tupla n,m, crear 3 conjuntos de datos distintos
# Por cada conjunto de datos, aplicar cada codigo
# Guardar el promedio de cada codigo al aplicarlo sobre los 3 casos posibles y almacenar valor. Asegurarse de ordenar la info en un nuevo excel ordenado

######################### Desarrolladores #########################
# Desarrollado por Diego, con colaboración de Benja
# Entrega los resultados necesarios para el desarrollo del proyecto 3 de ILN291
######################### Desarrolladores #########################

######################### Sobre el codigo #########################
# Codigo111: Utiliza el criterio p=1 para la generacion de solucion inicial, usa el primer criterio de generacion de vecinos y selecciona al que mas mejora de los vecinos
# Codigo112: Utiliza el criterio p=1 para la generacion de solucion inicial, usa el primer criterio de generacion de vecinos y selecciona al primero que mejora de los vecinos
# Codigo121: Utiliza el criterio p=1 para la generacion de solucion inicial, usa el segundo criterio de generacion de vecinos y selecciona al que mas mejora de los vecinos
# Codigo211: Utiliza el criterio p=10 para la generacion de solucion inicial, usa el primer criterio de generacion de vecinos y selecciona al que mas mejora de los vecinos
######################### Sobre el codigo #########################

######################### Criterios de generacion de vecinos #########################
# Criterio 1: Sobre cada ayudantia, cambia el estudiante asignado por otro, esto para cada ayudante. Especialidad: Si el estudiante al que se
# le asigna ya tenia una ayudantia asignada, entonces permuta las ayudantias entre ellos.
# Criterio 2: El mismo que el anterior, pero no existe permutacion. Este es mas sencillo ya que evita la permutacion, tambien permite encontrar
# nuevos vecinos que del otro criterio no son posibles, pero tiene como contra que genera mas casos infactibles. Por ende este codigo permite encontrar
# mas variedad de vecinos pero menos cantidad de vecinos factibles.
######################### Criterios de generacion de vecinos #########################

### Librerias ###

import random # Random parameters
import CodeLikeFunctions # Importants functions here !
import openpyxl # Work with Excel
import time # To time

### Codigo ###

# Seleccionamos los casos a estudiar y el número de casos para promediar

Valoresnm = [(70,70), (80,80), (90,90), (100,100), (110,110), (120,120), (130,130), (140,140), (150,150)]
N = 4

# Nombre Archivo

Nombre = "ResultadosDePruebaMuchosMuchosDatos.xlsx"

# Posibles horas de ayudantias
H_Ayudantia=[7,8,15]

# Creamos documento donde guardar datos

wb = openpyxl.Workbook()
tiempos = wb.create_sheet("Todos los tiempos")
valor_fo = wb.create_sheet("Valores FO")
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
tiempos_tabulados = wb.create_sheet("Tiempos en tablas")

## Detalles del excel

tiempos.cell(1,2).value = "Codigo111"
tiempos.cell(1,3).value = "Codigo112"
tiempos.cell(1,4).value = "Codigo121"
tiempos.cell(1,5).value = "Codigo211"
tiempos.cell(1,6).value = "Codigo222"
tiempos.cell(1,7).value = "Codigo221"

valor_fo.cell(1,2).value = "Codigo111"
valor_fo.cell(1,3).value = "Codigo112"
valor_fo.cell(1,4).value = "Codigo121"
valor_fo.cell(1,5).value = "Codigo211"
valor_fo.cell(1,6).value = "Codigo222"
valor_fo.cell(1,7).value = "Codigo221"

## Sobre cada tupla

h = 2
aux = 3
t = time.time()
for tupla in Valoresnm:
    n = tupla[0]; m = tupla[1]
    tiempos_tabulados.cell(aux, 1).value = str(n) + "n" + str(m) + "m"
    aux+=1
    # Creamos N casos sobre cada tupla
    for k in range(N):
        Horas = []
        Calificaciones = []
        # Creamos matriz Calificaciones
        for i in range(m):
            array_aux = []
            for j in range(n):
                array_aux.append(random.randint(0,10))
            Calificaciones.append(array_aux)
        # Creamos matriz horas de ayudantia
        for j in range(n):
            Horas.append(random.choice(H_Ayudantia))
        #print(Calificaciones, "\n", Horas)
        # Cabecera Excel
        tiempos.cell(h,1).value = str(n) + "n" + str(m) + "m"+ "(" + str(k) + ")"
        valor_fo.cell(h,1).value = str(n) + "n" + str(m) + "m"+ "(" + str(k) + ")"
        # Datos obtenido por cada caso
        #tiempos.cell(h,2).value, valor_fo.cell(h,2).value = CodeLikeFunctions.codigo111(n, m, Calificaciones, Horas)
        #tiempos.cell(h,3).value, valor_fo.cell(h,3).value = CodeLikeFunctions.codigo112(n, m, Calificaciones, Horas)
        #tiempos.cell(h,4).value, valor_fo.cell(h,4).value = CodeLikeFunctions.codigo121(n, m, Calificaciones, Horas)
        tiempos.cell(h,2).value, valor_fo.cell(h,2).value = -1, -1
        tiempos.cell(h,3).value, valor_fo.cell(h,3).value = -1, -1
        tiempos.cell(h,4).value, valor_fo.cell(h,4).value = -1, -1
        tiempos.cell(h,5).value, valor_fo.cell(h,5).value = CodeLikeFunctions.codigo211(n, m, Calificaciones, Horas)
        tiempos.cell(h,6).value, valor_fo.cell(h,6).value = CodeLikeFunctions.codigo222(n, m, Calificaciones, Horas)
        tiempos.cell(h,7).value, valor_fo.cell(h,7).value = CodeLikeFunctions.codigo221(n, m, Calificaciones, Horas)
        print("Caso "+ str(n) + "n" + str(m) + "m"+ "(" + str(k) + ")"+" listo, tiempo hasta ahora: "+ str(time.time()-t)+"\n")
        h+=1

## Tabulacion de datos resumidos

# Cabeceras
for i in range(6):
    tiempos_tabulados.cell(1,2*i+2).value = "tiempo"
    tiempos_tabulados.cell(1,2*i+3).value = "valor fo"
tiempos_tabulados.cell(2,2).value = "Codigo111"
tiempos_tabulados.cell(2,4).value = "Codigo112"
tiempos_tabulados.cell(2,6).value = "Codigo121"
tiempos_tabulados.cell(2,8).value = "Codigo211"
tiempos_tabulados.cell(2,10).value = "Codigo222"
tiempos_tabulados.cell(2,12).value = "Codigo221"

# Backup

wb.save(filename = Nombre)

# Datos promediados
for i in range(len(Valoresnm)):
    aux = [N,N,N,N,N,N]
    promedio_tiempo = [0,0,0,0,0,0]
    promedio_fo = [0,0,0,0,0,0]
    for j in range(N):
        for k in range(len(aux)):
            if float(tiempos.cell(N*i+2+j, k+2).value) == -1:
                aux[k] = aux[k] - 1
            else:
                promedio_tiempo[k] += float(tiempos.cell(N*i+2+j, k+2).value)
                promedio_fo[k] += float(valor_fo.cell(N*i+2+j, k+2).value)
    for j in range(len(aux)):
        if aux[j] <= 0:
            tiempos_tabulados.cell(3+i,2+2*j).value = -1
            tiempos_tabulados.cell(3+i,3+2*j).value = -1
        else:
            tiempos_tabulados.cell(3+i,2+2*j).value = promedio_tiempo[j]/aux[j]
            tiempos_tabulados.cell(3+i,3+2*j).value = promedio_fo[j]/aux[j]
wb.save(filename = Nombre)

        